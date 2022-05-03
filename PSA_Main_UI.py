import sys
import pandas as pd
import win32com.client
# from Regen_UI import regen_ui               # Load the function which runs a batch to regen the UI each time code runs.
# This is used for Development
from Read_Master import (Read_Master, Read_AnalyserStatus, Read_PeakControl, Read_VersionNumbers, Read_AnalyserIO, Read_A_08_Analyse, Read_PeakExtract,
                         Read_TempExtract, Read_A_17_Standard)  # import the function which reads from the Master PSA file
import datetime
import json
import base64
import glob
import openpyxl
from PSA_Home_Cal import Ui_PSAHome
from PSA_Page1 import Ui_PSAPage1
from PSA_Page2 import Ui_PSAPage2
from PSA_Page3 import Ui_PSAPage3
from PSA_Page4 import Ui_PSAPage4
from PSA_Page5 import Ui_PSAPage5
from PSA_Page6 import Ui_PSAPage6
from PyQt5.QtWidgets import (QApplication, QMainWindow, QFileDialog, QTableWidgetItem)
from PyQt5 import (QtCore, QtGui)
from PyQt5.QtCore import QDir
from dateutil.relativedelta import relativedelta

# import pyi_splash # this is just here for the packaging to allow the splash screen to close. It will always throw an error since the library cannot be installed.

# regen_ui()                  # Regenerate the UI. This is used to update the UI file after changes


AnalyserToProcess = 0
open_window = 0

# open the data interchange file as a global variable which can be accessed throughout the program.
# This will be accessed multiple times by various functions within the functions.

with open("C:\\PSAGen\\ReportData.json") as f:
    reportdata = json.load(f)                                                                   # Open a Blank JSON File


class HomeWindow(QMainWindow, Ui_PSAHome):
    psadatafname = ""                                                                           # Filename of the folder where the PSA report data is
    jsonin1fname = ""                                                                           # Filename of the previous Month's PSA Report JSON file
    jsonin2fname = ""                                                                           # Filename of the PSA Report JSON file from 2 Month's ago
    jsonoutfname = ""                                                                           # Filename of the output JSON file
    jsonname = ""                                                                               # Filename of relevant JSON File
    masterpsafname = "J:\\Client Analysers\\Analyser PSA Report\\PSA Report Config data.xlsx"   # Filename where the PSA Report Master configuration document is kept.
    PSAMasterData = ""                                                                          # Master file data variable from PSA Master file.
    AnalyserStatus = ""                                                                         # PSA Report file data from PSA Report file.
    PeakControl = ""                                                                            # Analyser Peak Control Data
    VersionNumbers = ""                                                                         # version numbers data
    AnalyserIO = ""                                                                             # Analyser IO Data Variable
    AnalyserA17 = ""                                                                            # Analyser Monthly Standardisation Data Variable
    AnalyserA08 = ""                                                                            # Analysis Data Variable
    PeakExtract = ""                                                                            # Peak Stability Data Variable
    TempExtract = ""                                                                            # Temp Stability Data Variable
    resourcepath = "C:\\PSAGen"                                                                 # Path to the UI Resources folder where all images are stored
    newanalyser = 0                                                                             # New Analyser flag
    autosend = 0                                                                                # Autosend Email Flag
    blankrep = 0                                                                                # Issue Blank report flag

    # pyi_splash.close()                                                                        # Used to close splash screen in packaged software

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        HomeWindow.PSAMasterData = Read_Master(HomeWindow.masterpsafname)  # Read in master file data from PSA Master file.
        self.page1 = PSA_pg1()
        self.page2 = PSA_pg2()
        self.page3 = PSA_pg3()
        self.page4 = PSA_pg4()
        self.page5 = PSA_pg5()
        self.page6 = PSA_pg6()
        self.connectSignalsSlots()
        self.popanalysers()
        self.psamasterloc.setText(HomeWindow.masterpsafname)  # Update PSA Master file location Label

    def connectSignalsSlots(self):
        self.PreviewReportPB.released.connect(self.updateReport)
        self.ImportReportPB.released.connect(self.importPSAData)
        self.ExportReportPB.released.connect(self.ExportReport)
        self.PB_pg.released.connect(self.show_page_1)
        self.PB_pg_2.released.connect(self.show_page_2)
        self.PB_pg_3.released.connect(self.show_page_3)
        self.PB_pg_4.released.connect(self.show_page_4)
        self.PB_pg_5.released.connect(self.show_page_5)
        self.PB_pg_6.released.connect(self.show_page_6)
        self.page1.NextPage.released.connect(self.UpdateJSON)
        self.page2.NextPage.released.connect(self.UpdateJSON)
        self.page6.RepAgo1.released.connect(self.NewAnUpdatePg6_1RepAgo)
        self.page6.RepAgo2.released.connect(self.NewAnUpdatePg6_2RepAgo)
        self.page6.NextPage.released.connect(self.UpdateJSON)
        self.NewAnalyser.stateChanged.connect(lambda: self.EnableUpdateHistData(self.NewAnalyser))
        self.AutoSendReport.stateChanged.connect(lambda: self.AutoSendToggle(self.AutoSendReport))
        self.AnalyserListComboBox.currentTextChanged.connect(self.AnalyserChanged)
        self.UpdateChecklistPB.released.connect(self.UpdatePSAChecklist)
        self.IssueBlankPB.released.connect(self.IssueBlank)

    def importPSAData(self):
        HomeWindow.psadatafname = self.GetFolder("Select Folder containing PSA Report Data")
        if HomeWindow.psadatafname:
            HomeWindow.PSAMasterData = Read_Master(HomeWindow.masterpsafname)  # Read in master file data from PSA Master file.
            print("Read Master Complete")
            HomeWindow.AnalyserStatus = Read_AnalyserStatus(HomeWindow.psadatafname)  # Read in PSA Report file data from PSA Report file.
            print("Read Analyser Status Complete")
            HomeWindow.PeakControl = Read_PeakControl(HomeWindow.psadatafname)  # Read in Analyser Peak Control Data
            print("Read Peak Control Complete")
            HomeWindow.VersionNumbers = Read_VersionNumbers(HomeWindow.psadatafname)  # Read in version numbers data
            print("Read Version Numbers Complete")
            HomeWindow.AnalyserIO = Read_AnalyserIO(HomeWindow.psadatafname)  # Read in Analyser IO Data
            print("Read Analyser IO Complete")
            HomeWindow.AnalyserA17 = Read_A_17_Standard(HomeWindow.psadatafname)  # Read in Analyser Standardisation data
            print("Read Analyser Std Data Complete")
            HomeWindow.AnalyserA08 = Read_A_08_Analyse(HomeWindow.psadatafname, HomeWindow.resourcepath, self.AnalyserListComboBox.currentText())  # Read in Analysis Data
            print("Read Analyser A_XX Complete")
            HomeWindow.PeakExtract = Read_PeakExtract(HomeWindow.psadatafname, HomeWindow.resourcepath)  # Readi in Peak Stability Data
            print("Read Peak Extract Complete")
            HomeWindow.TempExtract = Read_TempExtract(HomeWindow.psadatafname, HomeWindow.resourcepath)  # Read in Temp Stability Data
            print("Read Temp Extract Complete")
            self.PreviewReportPB.setEnabled(True)
            self.IssueBlankPB.setEnabled(False)
            self.psadataloc.setText(HomeWindow.psadatafname)
        if not HomeWindow.psadatafname:
            return

    def AutoSendToggle(self, b):
        # Toggles Auto Send Report Flag
        if b.isChecked():
            HomeWindow.autosend = 1

        else:
            HomeWindow.autosend = 0

    def EnableUpdateHistData(self, b):
        if b.isChecked():
            self.page6.RepAgo1.setEnabled(True)  # activate the import from 1 report Ago button pushbutton which will allow the data from 1 report ago to be updated.
            self.page6.RepAgo2.setEnabled(True)  # activate the import from 2 reports Ago button pushbutton which will allow the data from 2 reports ago to be updated.
        else:
            self.page6.RepAgo1.setEnabled(False)  # deactivate the import from 1 report Ago button pushbutton.
            self.page6.RepAgo2.setEnabled(False)  # deactivate the import from 2 reports Ago button pushbutton.

    def AnalyserChanged(self):  # determines if the Selected analyser has changed and if so disables the buttons
        self.PreviewReportPB.setEnabled(False)
        self.ExportReportPB.setEnabled(False)
        self.UpdateChecklistPB.setEnabled(False)
        self.ImportReportPB.setEnabled(True)
        self.PB_pg.setEnabled(False)
        self.PB_pg_2.setEnabled(False)
        self.PB_pg_3.setEnabled(False)
        self.PB_pg_4.setEnabled(False)
        self.PB_pg_5.setEnabled(False)
        self.PB_pg_6.setEnabled(False)
        HomeWindow.blankrep = 0

    def GetFolder(self, message):
        anal = self.AnalyserListComboBox.currentText()
        # Switches default folder based on which analyser is selected.
        if anal[0:3] == "C15":
            fname = QFileDialog.getExistingDirectory(self, message, "J:\\Client Analysers\\NG-1500")
        elif anal[0:3] == "C21":
            fname = QFileDialog.getExistingDirectory(self, message, "J:\\Client Analysers\\CS-2100")
        elif anal[0:3] == "CMM":
            fname = QFileDialog.getExistingDirectory(self, message, "J:\\Client Analysers\\CMM-100")
        elif anal[0:3] == "OBA":
            fname = QFileDialog.getExistingDirectory(self, message, "J:\\Client Analysers\\On Belt Analyser")
        elif anal[0:3] == "TBM":
            fname = QFileDialog.getExistingDirectory(self, message, "J:\\Client Analysers\\TBM-210")
        else:
            fname = QFileDialog.getExistingDirectory(self, message, "J:\\Client Analysers\\TBM-230")

        # ensures the file separators are correct for operating system
        if fname:
            fname = QDir.toNativeSeparators(fname)

        # if user presses cancel, return control to calling function
        if not fname:
            return

        return fname
        # Update the default path vairable in the ReadINData Class
        # HomeWindow.psadatafname = fname

        # # Update the data in the Data Variables in Use
        # self.updatereadindata()

    def JSONFname(self):

        # this function determines the filename of the output data file
        dates = HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "LastPsaReportTime", "Value"]  # get the date of the last psa report generation
        dates = pd.to_datetime(dates, yearfirst=True)  # setup dataframe as a date time series in the correct format
        mnth = dates.dt.month.to_string(index=False)
        yr = dates.dt.year.to_string(index=False)
        psafname = HomeWindow.psadatafname
        anal = self.page1.rep_analyser_data.text()

        if mnth == "10" or mnth == "11" or mnth == "12":
            fname = str(anal) + " PSA Report " + str(yr[2:4]) + str(mnth) + ".json"  # setting the filename which will be used to output the JSON file
        else:
            fname = str(anal) + " PSA Report " + str(yr[2:4]) + "0" + str(mnth) + ".json"  # setting the filename which will be used to output the JSON file
        HomeWindow.jsonname = fname
        HomeWindow.jsonoutfname = psafname + "\\" + HomeWindow.jsonname

    def UpdateJSON(self):

        self.JSONFname()  # Update Output JSON Filename
        print(HomeWindow.jsonoutfname)
        # Update JSON File
        reportdata['Summary'][0]['SiteName'] = self.page1.site_name.text()
        reportdata['Summary'][0]['ReportDate'] = self.page1.rep_date_data.text()
        reportdata['Summary'][0]['AnalyserNo'] = self.page1.rep_analyser_data.text()
        reportdata['Summary'][0]['ServEng'] = self.page1.rep_serv_eng_data.text()
        reportdata['Summary'][0]['Application'] = self.page1.rep_app_data.text()
        reportdata['Summary'][0]['Period'] = self.page1.period_data.text()
        reportdata['Summary'][0]['Email'] = self.page1.email_data.text()
        reportdata['Summary'][0]['NextPSA'] = str(self.page1.NextPSAMnth.currentText() + " " + self.page1.NextPSAYear.currentText())
        reportdata['Summary'][0]['TopUpDue'] = str(self.page1.TopUpMonth.currentText() + " " + self.page1.TopUpYear.currentText())

        # this loop populates the JSON by looking to see if any data exists in the first cell of the row. If not then it skips. If is does, the data is pulled into the JSON
        # pull the data from the Action Taken table row 1 to fill the JSON
        for i in range(0, 5):
            it = self.page1.ActionTakenTable.item(i, 0)
            if it and it.text():
                jsonid = "Action" + str(i + 1)
                reportdata['ActionTaken'][0][jsonid][0]['Date'] = str(self.page1.ActionTakenTable.item(i, 0).text())
                reportdata['ActionTaken'][0][jsonid][0]['Time'] = str(self.page1.ActionTakenTable.item(i, 1).text())
                reportdata['ActionTaken'][0][jsonid][0]['Action'] = str(self.page1.ActionTakenTable.item(i, 2).text())
                reportdata['ActionTaken'][0][jsonid][0]['Description'] = str(self.page1.ActionTakenTable.item(i, 3).text())
            else:
                pass

        # this loop populates the JSON by looking to see if any data exists in the first cell of the row. If not then it skips. If is does, the data is pulled into the JSON
        # Populating the Action required section of the JSON file
        for i in range(0, 5):
            it = self.page1.ActionTakenTable.item(i, 0)
            if it and it.text():
                jsonid = "ActionReq" + str(i + 1)
                reportdata['ActionRequired'][0][jsonid][0]['Action'] = str(self.page1.ActionRequiredTable.item(i, 0).text())
                reportdata['ActionRequired'][0][jsonid][0]['ByWhom'] = str(self.page1.ActionRequiredTable.item(i, 1).text())
                reportdata['ActionRequired'][0][jsonid][0]['ByWhen'] = str(self.page1.ActionRequiredTable.item(i, 2).text())
            else:
                pass

        # populating JSON with Detector stability info
        for i in range(0, 16):
            it = self.page2.tableWidget.item(0, i)
            if it and it.text():
                jsonid = "Detector" + str(i + 1)
                reportdata['DetectorStability'][0][jsonid][0]['Enabled'] = str(self.page2.tableWidget.item(0, i).text())
                reportdata['DetectorStability'][0][jsonid][0]['Stable'] = str(self.page2.tableWidget.item(1, i).text())

        # pull stability question data from the Temperature stability combo boxes
        reportdata['Temperatures'][0]['StableDetTemp'] = str(self.page2.DetTempStable.currentText())
        reportdata['Temperatures'][0]['StableElecTemp'] = str(self.page2.ElecCabTempStable.currentText())

        # this section dumps data from page 6 into the JSON.
        # PlantPLC Table dump to JSON
        it = self.page6.PlantPLCTable.item(0, 0)
        if it and it.text():
            reportdata['PLCStatus'][0]['BeltRunning'][0]['Current'] = str(self.page6.PlantPLCTable.item(0, 0).text())
            reportdata['PLCStatus'][0]['BeltRunning'][0]['1RepAgo'] = str(self.page6.PlantPLCTable.item(0, 1).text())
            reportdata['PLCStatus'][0]['BeltRunning'][0]['2RepAgo'] = str(self.page6.PlantPLCTable.item(0, 2).text())
            reportdata['PLCStatus'][0]['BeltRunning'][0]['Comment'] = str(self.page6.PlantPLCTable.item(0, 3).text())
            reportdata['PLCStatus'][0]['ForceAnalyse'][0]['Current'] = str(self.page6.PlantPLCTable.item(1, 0).text())
            reportdata['PLCStatus'][0]['ForceAnalyse'][0]['1RepAgo'] = str(self.page6.PlantPLCTable.item(1, 1).text())
            reportdata['PLCStatus'][0]['ForceAnalyse'][0]['2RepAgo'] = str(self.page6.PlantPLCTable.item(1, 2).text())
            reportdata['PLCStatus'][0]['ForceAnalyse'][0]['Comment'] = str(self.page6.PlantPLCTable.item(1, 3).text())
            reportdata['PLCStatus'][0]['ForceStandardise'][0]['Current'] = str(self.page6.PlantPLCTable.item(2, 0).text())
            reportdata['PLCStatus'][0]['ForceStandardise'][0]['1RepAgo'] = str(self.page6.PlantPLCTable.item(2, 1).text())
            reportdata['PLCStatus'][0]['ForceStandardise'][0]['2RepAgo'] = str(self.page6.PlantPLCTable.item(2, 2).text())
            reportdata['PLCStatus'][0]['ForceStandardise'][0]['Comment'] = str(self.page6.PlantPLCTable.item(2, 3).text())
        else:
            pass

        # AnalyserStatus Table dump to JSON
        it = self.page6.PlantPLCTable_2.item(0, 0)
        if it and it.text():
            reportdata['AnalyserStatus'][0]['AnalyserOK'][0]['Current'] = str(self.page6.PlantPLCTable_2.item(0, 0).text())
            reportdata['AnalyserStatus'][0]['AnalyserOK'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_2.item(0, 1).text())
            reportdata['AnalyserStatus'][0]['AnalyserOK'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_2.item(0, 2).text())
            reportdata['AnalyserStatus'][0]['AnalyserOK'][0]['Comment'] = str(self.page6.PlantPLCTable_2.item(0, 3).text())
            reportdata['AnalyserStatus'][0]['StandardsOK'][0]['Current'] = str(self.page6.PlantPLCTable_2.item(1, 0).text())
            reportdata['AnalyserStatus'][0]['StandardsOK'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_2.item(1, 1).text())
            reportdata['AnalyserStatus'][0]['StandardsOK'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_2.item(1, 2).text())
            reportdata['AnalyserStatus'][0]['StandardsOK'][0]['Comment'] = str(self.page6.PlantPLCTable_2.item(1, 3).text())
            reportdata['AnalyserStatus'][0]['IOControlOK'][0]['Current'] = str(self.page6.PlantPLCTable_2.item(2, 0).text())
            reportdata['AnalyserStatus'][0]['IOControlOK'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_2.item(2, 1).text())
            reportdata['AnalyserStatus'][0]['IOControlOK'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_2.item(2, 2).text())
            reportdata['AnalyserStatus'][0]['IOControlOK'][0]['Comment'] = str(self.page6.PlantPLCTable_2.item(2, 3).text())
            reportdata['AnalyserStatus'][0]['SpectraStable'][0]['Current'] = str(self.page6.PlantPLCTable_2.item(3, 0).text())
            reportdata['AnalyserStatus'][0]['SpectraStable'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_2.item(3, 1).text())
            reportdata['AnalyserStatus'][0]['SpectraStable'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_2.item(3, 2).text())
            reportdata['AnalyserStatus'][0]['SpectraStable'][0]['Comment'] = str(self.page6.PlantPLCTable_2.item(3, 3).text())
        else:
            pass

        # PLC Analyser Results Table dump to JSON
        it = self.page6.PlantPLCTable_3.item(0, 0)
        if it and it.text():
            reportdata['PLCResults'][0]['SystemOK'][0]['Current'] = str(self.page6.PlantPLCTable_3.item(0, 0).text())
            reportdata['PLCResults'][0]['SystemOK'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_3.item(0, 1).text())
            reportdata['PLCResults'][0]['SystemOK'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_3.item(0, 2).text())
            reportdata['PLCResults'][0]['SystemOK'][0]['Comment'] = str(self.page6.PlantPLCTable_3.item(0, 3).text())
            reportdata['PLCResults'][0]['SourceControlFault'][0]['Current'] = str(self.page6.PlantPLCTable_3.item(1, 0).text())
            reportdata['PLCResults'][0]['SourceControlFault'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_3.item(1, 1).text())
            reportdata['PLCResults'][0]['SourceControlFault'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_3.item(1, 2).text())
            reportdata['PLCResults'][0]['SourceControlFault'][0]['Comment'] = str(self.page6.PlantPLCTable_3.item(1, 3).text())
            reportdata['PLCResults'][0]['SourceOff'][0]['Current'] = str(self.page6.PlantPLCTable_3.item(2, 0).text())
            reportdata['PLCResults'][0]['SourceOff'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_3.item(2, 1).text())
            reportdata['PLCResults'][0]['SourceOff'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_3.item(2, 2).text())
            reportdata['PLCResults'][0]['SourceOff'][0]['Comment'] = str(self.page6.PlantPLCTable_3.item(2, 3).text())
            reportdata['PLCResults'][0]['SourceOn'][0]['Current'] = str(self.page6.PlantPLCTable_3.item(3, 0).text())
            reportdata['PLCResults'][0]['SourceOn'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_3.item(3, 1).text())
            reportdata['PLCResults'][0]['SourceOn'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_3.item(3, 2).text())
            reportdata['PLCResults'][0]['SourceOn'][0]['Comment'] = str(self.page6.PlantPLCTable_3.item(3, 3).text())
        else:
            pass

        # Analyser Configuration Table dump to JSON
        it = self.page6.PlantPLCTable_4.item(0, 0)
        if it and it.text():
            reportdata['AnalyserConfiguration'][0]['AnalysisPeriod'][0]['Current'] = str(self.page6.PlantPLCTable_4.item(0, 0).text())
            reportdata['AnalyserConfiguration'][0]['AnalysisPeriod'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_4.item(0, 1).text())
            reportdata['AnalyserConfiguration'][0]['AnalysisPeriod'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_4.item(0, 2).text())
            reportdata['AnalyserConfiguration'][0]['AnalysisPeriod'][0]['Comment'] = str(self.page6.PlantPLCTable_4.item(0, 3).text())
            reportdata['AnalyserConfiguration'][0]['AnalMinLoadLimit'][0]['Current'] = str(self.page6.PlantPLCTable_4.item(1, 0).text())
            reportdata['AnalyserConfiguration'][0]['AnalMinLoadLimit'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_4.item(1, 1).text())
            reportdata['AnalyserConfiguration'][0]['AnalMinLoadLimit'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_4.item(1, 2).text())
            reportdata['AnalyserConfiguration'][0]['AnalMinLoadLimit'][0]['Comment'] = str(self.page6.PlantPLCTable_4.item(1, 3).text())
            reportdata['AnalyserConfiguration'][0]['StandardisePeriod'][0]['Current'] = str(self.page6.PlantPLCTable_4.item(2, 0).text())
            reportdata['AnalyserConfiguration'][0]['StandardisePeriod'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_4.item(2, 1).text())
            reportdata['AnalyserConfiguration'][0]['StandardisePeriod'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_4.item(2, 2).text())
            reportdata['AnalyserConfiguration'][0]['StandardisePeriod'][0]['Comment'] = str(self.page6.PlantPLCTable_4.item(2, 3).text())
        else:
            pass

        # Standardisation Table dump to JSON
        it = self.page6.PlantPLCTable_5.item(0, 0)
        if it and it.text():
            reportdata['Standardisation'][0]['FirstStandardTime'][0]['Current'] = str(self.page6.PlantPLCTable_5.item(0, 0).text())
            reportdata['Standardisation'][0]['FirstStandardTime'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_5.item(0, 1).text())
            reportdata['Standardisation'][0]['FirstStandardTime'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_5.item(0, 2).text())
            reportdata['Standardisation'][0]['FirstStandardTime'][0]['Comment'] = str(self.page6.PlantPLCTable_5.item(0, 3).text())
            reportdata['Standardisation'][0]['MostRecentStandard'][0]['Current'] = str(self.page6.PlantPLCTable_5.item(1, 0).text())
            reportdata['Standardisation'][0]['MostRecentStandard'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_5.item(1, 1).text())
            reportdata['Standardisation'][0]['MostRecentStandard'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_5.item(1, 2).text())
            reportdata['Standardisation'][0]['MostRecentStandard'][0]['Comment'] = str(self.page6.PlantPLCTable_5.item(1, 3).text())
            reportdata['Standardisation'][0]['NumStdPeriodsSinceClear'][0]['Current'] = str(self.page6.PlantPLCTable_5.item(2, 0).text())
            reportdata['Standardisation'][0]['NumStdPeriodsSinceClear'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_5.item(2, 1).text())
            reportdata['Standardisation'][0]['NumStdPeriodsSinceClear'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_5.item(2, 2).text())
            reportdata['Standardisation'][0]['NumStdPeriodsSinceClear'][0]['Comment'] = str(self.page6.PlantPLCTable_5.item(2, 3).text())
            reportdata['Standardisation'][0]['NumStdPeriodsThisMnth'][0]['Current'] = str(self.page6.PlantPLCTable_5.item(3, 0).text())
            reportdata['Standardisation'][0]['NumStdPeriodsThisMnth'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_5.item(3, 1).text())
            reportdata['Standardisation'][0]['NumStdPeriodsThisMnth'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_5.item(3, 2).text())
            reportdata['Standardisation'][0]['NumStdPeriodsThisMnth'][0]['Comment'] = str(self.page6.PlantPLCTable_5.item(3, 3).text())
        else:
            pass

        # SSoftware Versions Table dump to JSON
        it = self.page6.PlantPLCTable_6.item(0, 0)
        if it and it.text():
            reportdata['SoftwareVersions'][0]['Product'][0]['Current'] = str(self.page6.PlantPLCTable_6.item(0, 0).text())
            reportdata['SoftwareVersions'][0]['Product'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_6.item(0, 1).text())
            reportdata['SoftwareVersions'][0]['Product'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_6.item(0, 2).text())
            reportdata['SoftwareVersions'][0]['Product'][0]['Comment'] = str(self.page6.PlantPLCTable_6.item(0, 3).text())
            reportdata['SoftwareVersions'][0]['CsSchedule'][0]['Current'] = str(self.page6.PlantPLCTable_6.item(1, 0).text())
            reportdata['SoftwareVersions'][0]['CsSchedule'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_6.item(1, 1).text())
            reportdata['SoftwareVersions'][0]['CsSchedule'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_6.item(1, 2).text())
            reportdata['SoftwareVersions'][0]['CsSchedule'][0]['Comment'] = str(self.page6.PlantPLCTable_6.item(1, 3).text())
        else:
            pass

        it = self.page6.PlantPLCTable_7.item(0, 0)
        if it and it.text():
            reportdata['DiskSpaceMem'][0]['DiskSpace'][0]['Current'] = str(self.page6.PlantPLCTable_7.item(0, 0).text())
            reportdata['DiskSpaceMem'][0]['DiskSpace'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_7.item(0, 1).text())
            reportdata['DiskSpaceMem'][0]['DiskSpace'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_7.item(0, 2).text())
            reportdata['DiskSpaceMem'][0]['DiskSpace'][0]['Comment'] = str(self.page6.PlantPLCTable_7.item(0, 3).text())
            reportdata['DiskSpaceMem'][0]['PercDiskSpace'][0]['Current'] = str(self.page6.PlantPLCTable_7.item(1, 0).text())
            reportdata['DiskSpaceMem'][0]['PercDiskSpace'][0]['1RepAgo'] = str(self.page6.PlantPLCTable_7.item(1, 1).text())
            reportdata['DiskSpaceMem'][0]['PercDiskSpace'][0]['2RepAgo'] = str(self.page6.PlantPLCTable_7.item(1, 2).text())
            reportdata['DiskSpaceMem'][0]['PercDiskSpace'][0]['Comment'] = str(self.page6.PlantPLCTable_7.item(1, 3).text())
        else:
            pass

        # This section encodes the plots into the JSON File for transmission to engineer
        # encode Temperature Plot for storage in JSON
        with open(HomeWindow.resourcepath + "\\Temperatures.png", mode="rb") as tf:
            data = base64.b64encode(tf.read()).decode("utf-8")
        reportdata['Plots'][0]['TempPlot'] = data
        with open(HomeWindow.jsonoutfname, 'w') as file:
            json.dump(reportdata, file, indent=1)

        # encode Detector Stability Plot for storage in JSON
        with open(HomeWindow.resourcepath + "\\Detector_Stability.png", mode="rb") as tf:
            data = base64.b64encode(tf.read()).decode("utf-8")
        reportdata['Plots'][0]['DetStabPlot'] = data
        with open(HomeWindow.jsonoutfname, 'w') as file:
            json.dump(reportdata, file, indent=1)

        # encode Daily Tonnes Plot for storage in JSON
        with open(HomeWindow.resourcepath + "\\Daily_Tonnes.png", mode="rb") as tf:
            data = base64.b64encode(tf.read()).decode("utf-8")
        reportdata['Plots'][0]['DailyTonsPlot'] = data
        with open(HomeWindow.jsonoutfname, 'w') as file:
            json.dump(reportdata, file, indent=1)

        # encode Results 1 Plot for storage in JSON
        with open(HomeWindow.resourcepath + "\\Results1.png", mode="rb") as tf:
            data = base64.b64encode(tf.read()).decode("utf-8")
        reportdata['Plots'][0]['Results1Plot'] = data
        with open(HomeWindow.jsonoutfname, 'w') as file:
            json.dump(reportdata, file, indent=1)

        # encode Results 2 Plot for storage in JSON
        with open(HomeWindow.resourcepath + "\\Results2.png", mode="rb") as tf:
            data = base64.b64encode(tf.read()).decode("utf-8")
        reportdata['Plots'][0]['Results2Plot'] = data
        with open(HomeWindow.jsonoutfname, 'w') as file:
            json.dump(reportdata, file, indent=1)

        print("JSON Updated")
        self.ExportReportPB.setEnabled(True)

    def ExportReport(self):
        with open(HomeWindow.jsonoutfname) as f:
            jsondata = json.load(f)
        configdatafname = "J:\\Client Analysers\\Analyser PSA Report\\PSA Report Config data.xlsx"  # generates the checklist filename based on the year of the period of the report being produced. This will deal with the January / december issue
        configdata = openpyxl.load_workbook(configdatafname)  # Load config data from Config Data spreadsheet
        engineers = configdata["Engineers"]  # Select Engineers sheet
        serveng = jsondata["Summary"][0]["ServEng"]  # load relevant service engineers from JSON
        maxrow = engineers.max_row  # Determine max row of document

        for row in range(1, maxrow + 1):  # For loop for searching through the engineers sheet to determine the email address
            engcol = "{}{}".format("A", row)  # this sets up the cell reference in the PSA report Config Excel document that we are going to search through.
            emailcol = "{}{}".format("B", row)  # this sets up the cell reference in the PSA report Config Excel document that we are going to find the email address.
            engnamecol = "{}{}".format("C", row)  # this sets up the cell reference in the PSA report Config Excel document that we are going to find the engineer's preferred name.
            x = engineers[engcol].value  # Load the engineers name into a variable for checking

            if x == serveng:  # if statement that triggers when the engineers name is found.
                engemail = engineers[emailcol].value  # put the engineers email address into the engemail variable which is used later for setting the engineers email address in the email
                engname = engineers[engnamecol].value  # put the engineer's preferred name in the variable engname
                break
        outlook = win32com.client.Dispatch('outlook.application')  # invoke outlook to automatically generate the email
        mail = outlook.CreateItem(0)  # This creates the email object. The 0 refers to the item type from the office documentation OlItemType. This can be others eg 1 for appointments, 2 for contacts etc.
        mail.To = engemail  # insert the relevant engineer's email address
        mail.CC = 'l.biggins@scantech.com.au; i.nerush@scantech.com.au; d.rossouw@scantech.com.au; l.balzan@scantech.com.au; m.kalicinski@scantech.com.au'
        if HomeWindow.blankrep == 1:
            subjectline = str(jsondata["Summary"][0]["AnalyserNo"]) + " Blank Monthly PSA Report data for " + str(jsondata["Summary"][0]["Period"])
            emailbody = "<html><body style=font-family:Calibri;> Hi " + str(engname) + " <br><br> Attached is the monthly report data " \
                                                                                       "for " + str(jsondata["Summary"][0]["AnalyserNo"]) + ".<br>No PSA data was received this month. This was due " \
                                                                                                                                            "to either the analyser or the remote connection being offline.<br>Please look into the issue and see if it can be " \
                                                                                                                                            "resolved for next month's report. <br><br>Even though this report is blank, there should be an attachment " \
                                                                                                                                            "called " + str(
                HomeWindow.jsonname) + " which contains the blank report information you will need to generate the blank " \
                                       "pdf report. <br><br> </body></html>"
        else:
            subjectline = str(jsondata["Summary"][0]["AnalyserNo"]) + " Monthly PSA Report data for " + str(jsondata["Summary"][0]["Period"])
            emailbody = "<html><body style=font-family:Calibri;> Hi " + str(engname) + " <br><br> Attached is the monthly report data " \
                                                                                       "for " + str(jsondata["Summary"][0]["AnalyserNo"]) + ". <br>You will need this JSON file to generate your PDF report." \
                                                                                                                                            "<br>Please remember to CC Igor, Lucas & Daniel when sending your PDF report and upload the PDF and JSON, overwriting" \
                                                                                                                                            " the old JSON, to the server once sent to the customer."
        mail.Subject = subjectline  # Build the subject line automatically.
        mail.BodyFormat = 2
        mail.HTMLBody = emailbody
        mail.Attachments.Add(HomeWindow.jsonoutfname)  # add the json file as an attachment
        if HomeWindow.autosend == 0:  # if statement to determine if the email should be sent directly or displayed for manual sending.
            mail.Display()
        elif HomeWindow.autosend == 1:
            mail.Send()
        print("Exporting report to email")
        self.UpdateChecklistPB.setEnabled(True)  # activate the update checklist pushbutton which will allow the PSA report checklist to be updated.

    def updateReport(self):

        self.ExportReportPB.setEnabled(False)
        self.PB_pg.setEnabled(True)
        self.PB_pg_2.setEnabled(True)
        self.PB_pg_3.setEnabled(True)
        self.PB_pg_4.setEnabled(True)
        self.PB_pg_5.setEnabled(True)
        self.PB_pg_6.setEnabled(True)
        # Update data accordingly
        anal = self.AnalyserListComboBox.currentText()
        repdate = str(datetime.date.today())  # get today's date
        serveng = HomeWindow.PSAMasterData.loc[HomeWindow.PSAMasterData["Analyser Number"] == anal, "Service Engineer"].to_string(index=False)  # get the service engineer as a string without dataframe nonsense
        application = HomeWindow.PSAMasterData.loc[HomeWindow.PSAMasterData["Analyser Number"] == anal, "Application"].to_string(index=False)  # get the service engineer as a string without dataframe nonsense
        customer = HomeWindow.PSAMasterData.loc[HomeWindow.PSAMasterData["Analyser Number"] == anal, "Customer Name"].to_string(index=False)  # get the service engineer as a string without dataframe nonsense
        region = HomeWindow.PSAMasterData.loc[HomeWindow.PSAMasterData["Analyser Number"] == anal, "Region"].to_string(index=False)  # get the service engineer as a string without dataframe nonsense
        # Update UI
        self.page1.rep_analyser_data.setText(anal)
        self.page1.rep_date_data.setText(repdate)
        self.page1.rep_serv_eng_data.setText(serveng)
        self.page1.rep_app_data.setText(application)
        self.page1.site_name.setText(customer)
        self.page1.period_data.setText(self.RepPeriod())
        self.page1.email_data.setText(self.RepEmail(region))
        self.page1.disp_stduptodate.setText(self.StdDate())
        self.page1.disp_endiskspc.setText(self.DiskSpaceOK())
        self.UpdateDetStab()
        self.UpdatePg6()
        self.UpdatePg1()
        self.UpdateFigs()
        self.show_page_1()
        print("Report Updated")

    def EnabledYes(self):
        enabled = "Yes"
        enableditem = QTableWidgetItem(enabled)
        brush = QtGui.QBrush(QtGui.QColor(85, 255, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        enableditem.setBackground(brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        enableditem.setForeground(brush)
        return enableditem

    def EnabledNo(self):
        enabled = "No"
        enableditem = QTableWidgetItem(enabled)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        enableditem.setBackground(brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        enableditem.setForeground(brush)
        return enableditem

    def ClearTable(self):
        enabled = ""
        enableditem = QTableWidgetItem(enabled)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        enableditem.setBackground(brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        enableditem.setForeground(brush)
        return enableditem

    def UpdatePg1(self):
        fname = HomeWindow.jsonin1fname
        with open(fname) as f:  # open json using filename generated earlier.
            reportdata = json.load(f)
        # Update the Action Taken Table on Page 1
        self.page1.ActionTakenTable.setItem(0, 0, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action1'][0]['Date'])))
        self.page1.ActionTakenTable.setItem(0, 1, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action1'][0]['Time'])))
        self.page1.ActionTakenTable.setItem(0, 2, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action1'][0]['Action'])))
        self.page1.ActionTakenTable.setItem(0, 3, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action1'][0]['Description'])))
        self.page1.ActionTakenTable.setItem(1, 0, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action2'][0]['Date'])))
        self.page1.ActionTakenTable.setItem(1, 1, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action2'][0]['Time'])))
        self.page1.ActionTakenTable.setItem(1, 2, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action2'][0]['Action'])))
        self.page1.ActionTakenTable.setItem(1, 3, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action2'][0]['Description'])))
        self.page1.ActionTakenTable.setItem(2, 0, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action3'][0]['Date'])))
        self.page1.ActionTakenTable.setItem(2, 1, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action3'][0]['Time'])))
        self.page1.ActionTakenTable.setItem(2, 2, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action3'][0]['Action'])))
        self.page1.ActionTakenTable.setItem(2, 3, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action3'][0]['Description'])))
        self.page1.ActionTakenTable.setItem(3, 0, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action4'][0]['Date'])))
        self.page1.ActionTakenTable.setItem(3, 1, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action4'][0]['Time'])))
        self.page1.ActionTakenTable.setItem(3, 2, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action4'][0]['Action'])))
        self.page1.ActionTakenTable.setItem(3, 3, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action4'][0]['Description'])))
        self.page1.ActionTakenTable.setItem(4, 0, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action5'][0]['Date'])))
        self.page1.ActionTakenTable.setItem(4, 1, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action5'][0]['Time'])))
        self.page1.ActionTakenTable.setItem(4, 2, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action5'][0]['Action'])))
        self.page1.ActionTakenTable.setItem(4, 3, QTableWidgetItem(str(reportdata['ActionTaken'][0]['Action5'][0]['Description'])))
        # Update the Action required table on page 1
        self.page1.ActionRequiredTable.setItem(0, 0, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq1'][0]['Action'])))
        self.page1.ActionRequiredTable.setItem(0, 1, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq1'][0]['ByWhom'])))
        self.page1.ActionRequiredTable.setItem(0, 2, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq1'][0]['ByWhen'])))
        self.page1.ActionRequiredTable.setItem(1, 0, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq2'][0]['Action'])))
        self.page1.ActionRequiredTable.setItem(1, 1, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq2'][0]['ByWhom'])))
        self.page1.ActionRequiredTable.setItem(1, 2, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq2'][0]['ByWhen'])))
        self.page1.ActionRequiredTable.setItem(2, 0, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq3'][0]['Action'])))
        self.page1.ActionRequiredTable.setItem(2, 1, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq3'][0]['ByWhom'])))
        self.page1.ActionRequiredTable.setItem(2, 2, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq3'][0]['ByWhen'])))
        self.page1.ActionRequiredTable.setItem(3, 0, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq4'][0]['Action'])))
        self.page1.ActionRequiredTable.setItem(3, 1, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq4'][0]['ByWhom'])))
        self.page1.ActionRequiredTable.setItem(3, 2, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq4'][0]['ByWhen'])))
        self.page1.ActionRequiredTable.setItem(4, 0, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq5'][0]['Action'])))
        self.page1.ActionRequiredTable.setItem(4, 1, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq5'][0]['ByWhom'])))
        self.page1.ActionRequiredTable.setItem(4, 2, QTableWidgetItem(str(reportdata['ActionRequired'][0]['ActionReq5'][0]['ByWhen'])))

    def UpdatePg6(self):

        if self.IncLastMonth.isChecked():  # If statement that checks the status of the include last month checkbox. If checked will ask user for folder where last report's json is else uses default.
            fname1 = self.GetFolder("Select Folder containing Last Month's PSA Report Data")  # gets the filename of the JSON file containing last month's data
            if fname1:  # if statement to execute default file load in the event the user cancels the previous report load
                fname = glob.glob(fname1 + "\*.json")  # if the user has selected a folder, we then need to check if a json file exists in that folder
                if fname:  # if the json exists in the folder then generate the correct file name to allow it to be opened
                    fname = QDir.toNativeSeparators(fname[0])
                    HomeWindow.jsonin1fname = fname
                else:  # if the json doesn't exist in this folder, print json not available and use default
                    print("JSON not available in " + fname1)
                    print("Using blank default JSON in C:\\PSAGen")
                    fname = "C:\\PSAGen\\ReportData.json"
            if not fname1:  # if user cancels, no file name will be produced so print error and use default.
                print("Last report folder not selected, using default instead")
                print("Using blank default JSON in C:\\PSAGen")
                fname = "C:\\PSAGen\\ReportData.json"

            with open(fname) as f:  # open json using filename generated earlier.
                reportdata = json.load(f)
        else:
            print("Excluding last report, so using blank defaults in C:\\PSAGen instead")
            with open("C:\\PSAGen\\ReportData.json") as f:  # force default in event include last month is unchecked.
                reportdata = json.load(f)

        # Update PLC Status Table
        BeltRunning = QTableWidgetItem(HomeWindow.AnalyserIO.loc[HomeWindow.AnalyserIO["Parameter Name"] == "BeltRunning", "Value"].to_string(index=False))
        ForceAnalyse = QTableWidgetItem(HomeWindow.AnalyserIO.loc[HomeWindow.AnalyserIO["Parameter Name"] == "ForceAnalyse", "Value"].to_string(index=False))
        ForceStandardise = QTableWidgetItem(HomeWindow.AnalyserIO.loc[HomeWindow.AnalyserIO["Parameter Name"] == "ForceStandardise", "Value"].to_string(index=False))

        self.page6.PlantPLCTable.setItem(0, 0, BeltRunning)
        self.page6.PlantPLCTable.setItem(1, 0, ForceAnalyse)
        self.page6.PlantPLCTable.setItem(2, 0, ForceStandardise)
        # Creating the comment item and populating with Blanks
        self.page6.PlantPLCTable.setItem(0, 1, QTableWidgetItem(str(reportdata['PLCStatus'][0]['BeltRunning'][0]['Current'])))
        self.page6.PlantPLCTable.setItem(1, 1, QTableWidgetItem(str(reportdata['PLCStatus'][0]['ForceAnalyse'][0]['Current'])))
        self.page6.PlantPLCTable.setItem(2, 1, QTableWidgetItem(str(reportdata['PLCStatus'][0]['ForceStandardise'][0]['Current'])))
        self.page6.PlantPLCTable.setItem(0, 2, QTableWidgetItem(str(reportdata['PLCStatus'][0]['BeltRunning'][0]['1RepAgo'])))
        self.page6.PlantPLCTable.setItem(1, 2, QTableWidgetItem(str(reportdata['PLCStatus'][0]['ForceAnalyse'][0]['1RepAgo'])))
        self.page6.PlantPLCTable.setItem(2, 2, QTableWidgetItem(str(reportdata['PLCStatus'][0]['ForceStandardise'][0]['1RepAgo'])))
        self.page6.PlantPLCTable.setItem(0, 3, QTableWidgetItem(str(reportdata['PLCStatus'][0]['BeltRunning'][0]['Comment'])))
        self.page6.PlantPLCTable.setItem(1, 3, QTableWidgetItem(str(reportdata['PLCStatus'][0]['ForceAnalyse'][0]['Comment'])))
        self.page6.PlantPLCTable.setItem(2, 3, QTableWidgetItem(str(reportdata['PLCStatus'][0]['ForceStandardise'][0]['Comment'])))
        # Update Analyser Status Table
        AnalyserOK = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "AnalyserOK", "Value"].to_string(index=False))
        StandardsOK = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "StandardsOK", "Value"].to_string(index=False))
        IOControlOK = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "IOControlOK", "Value"].to_string(index=False))
        SpectraStable = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "SpectraStable", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_2.setItem(0, 0, AnalyserOK)
        self.page6.PlantPLCTable_2.setItem(1, 0, StandardsOK)
        self.page6.PlantPLCTable_2.setItem(2, 0, IOControlOK)
        self.page6.PlantPLCTable_2.setItem(3, 0, SpectraStable)
        # Creating the comment item and populating with Blanks
        self.page6.PlantPLCTable_2.setItem(0, 1, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['AnalyserOK'][0]['Current'])))
        self.page6.PlantPLCTable_2.setItem(1, 1, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['StandardsOK'][0]['Current'])))
        self.page6.PlantPLCTable_2.setItem(2, 1, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['IOControlOK'][0]['Current'])))
        self.page6.PlantPLCTable_2.setItem(3, 1, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['SpectraStable'][0]['Current'])))
        self.page6.PlantPLCTable_2.setItem(0, 2, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['AnalyserOK'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_2.setItem(1, 2, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['StandardsOK'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_2.setItem(2, 2, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['IOControlOK'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_2.setItem(3, 2, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['SpectraStable'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_2.setItem(0, 3, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['AnalyserOK'][0]['Comment'])))
        self.page6.PlantPLCTable_2.setItem(1, 3, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['StandardsOK'][0]['Comment'])))
        self.page6.PlantPLCTable_2.setItem(2, 3, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['IOControlOK'][0]['Comment'])))
        self.page6.PlantPLCTable_2.setItem(3, 3, QTableWidgetItem(str(reportdata['AnalyserStatus'][0]['SpectraStable'][0]['Comment'])))

        # Update PLC Analyser Results
        SystemOK = QTableWidgetItem(HomeWindow.AnalyserIO.loc[HomeWindow.AnalyserIO["Parameter Name"] == "SystemRunning", "Value"].to_string(index=False))
        if HomeWindow.AnalyserIO.loc[HomeWindow.AnalyserIO["Parameter Name"] == "SourceControlFault", "Value"].to_string(index=False) == "   OK\nFALSE":
            SourceControlFault = QTableWidgetItem("FALSE")
        else:
            SourceControlFault = QTableWidgetItem("TRUE")
        SourceOff = QTableWidgetItem(HomeWindow.AnalyserIO.loc[HomeWindow.AnalyserIO["Parameter Name"] == "SourceOffProx", "Value"].to_string(index=False))
        SourceOn = QTableWidgetItem(HomeWindow.AnalyserIO.loc[HomeWindow.AnalyserIO["Parameter Name"] == "SourceOnProx", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_3.setItem(0, 0, SystemOK)
        self.page6.PlantPLCTable_3.setItem(1, 0, SourceControlFault)
        self.page6.PlantPLCTable_3.setItem(2, 0, SourceOff)
        self.page6.PlantPLCTable_3.setItem(3, 0, SourceOn)
        # Creating the comment item and populating with Blanks
        self.page6.PlantPLCTable_3.setItem(0, 1, QTableWidgetItem(str(reportdata['PLCResults'][0]['SystemOK'][0]['Current'])))
        self.page6.PlantPLCTable_3.setItem(1, 1, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceControlFault'][0]['Current'])))
        self.page6.PlantPLCTable_3.setItem(2, 1, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceOff'][0]['Current'])))
        self.page6.PlantPLCTable_3.setItem(3, 1, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceOn'][0]['Current'])))
        self.page6.PlantPLCTable_3.setItem(0, 2, QTableWidgetItem(str(reportdata['PLCResults'][0]['SystemOK'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_3.setItem(1, 2, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceControlFault'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_3.setItem(2, 2, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceOff'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_3.setItem(3, 2, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceOn'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_3.setItem(0, 3, QTableWidgetItem(str(reportdata['PLCResults'][0]['SystemOK'][0]['Comment'])))
        self.page6.PlantPLCTable_3.setItem(1, 3, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceControlFault'][0]['Comment'])))
        self.page6.PlantPLCTable_3.setItem(2, 3, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceOff'][0]['Comment'])))
        self.page6.PlantPLCTable_3.setItem(3, 3, QTableWidgetItem(str(reportdata['PLCResults'][0]['SourceOn'][0]['Comment'])))

        # Update Analyser Config Table
        AnalysisPeriod = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "AnalysisPeriod", "Value"].to_string(index=False))
        AnalMinLoadLimit = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "AnalMinLoadLimit", "Value"].to_string(index=False))
        StandardisePeriod = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "StandardisePeriod", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_4.setItem(0, 0, AnalysisPeriod)
        self.page6.PlantPLCTable_4.setItem(1, 0, AnalMinLoadLimit)
        self.page6.PlantPLCTable_4.setItem(2, 0, StandardisePeriod)
        # Creating the comment item and populating with Blanks
        self.page6.PlantPLCTable_4.setItem(0, 1, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['AnalysisPeriod'][0]['Current'])))
        self.page6.PlantPLCTable_4.setItem(1, 1, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['AnalMinLoadLimit'][0]['Current'])))
        self.page6.PlantPLCTable_4.setItem(2, 1, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['StandardisePeriod'][0]['Current'])))
        self.page6.PlantPLCTable_4.setItem(0, 2, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['AnalysisPeriod'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_4.setItem(1, 2, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['AnalMinLoadLimit'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_4.setItem(2, 2, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['StandardisePeriod'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_4.setItem(0, 3, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['AnalysisPeriod'][0]['Comment'])))
        self.page6.PlantPLCTable_4.setItem(1, 3, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['AnalMinLoadLimit'][0]['Comment'])))
        self.page6.PlantPLCTable_4.setItem(2, 3, QTableWidgetItem(str(reportdata['AnalyserConfiguration'][0]['StandardisePeriod'][0]['Comment'])))

        # Update Standardisation Table
        FirstStandard = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "FirstStandardTime", "Value"].to_string(index=False))
        LastStandard = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "LastStandardiseTime", "Value"].to_string(index=False))
        NumStandard = QTableWidgetItem(HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "PeriodCount", "Value"].to_string(index=False))
        NumStandardMonth = QTableWidgetItem(str(HomeWindow.AnalyserA17))
        self.page6.PlantPLCTable_5.setItem(0, 0, FirstStandard)
        self.page6.PlantPLCTable_5.setItem(1, 0, LastStandard)
        self.page6.PlantPLCTable_5.setItem(2, 0, NumStandard)
        self.page6.PlantPLCTable_5.setItem(3, 0, NumStandardMonth)
        # Creating the comment item and populating with Blanks
        self.page6.PlantPLCTable_5.setItem(0, 1, QTableWidgetItem(str(reportdata['Standardisation'][0]['FirstStandardTime'][0]['Current'])))
        self.page6.PlantPLCTable_5.setItem(1, 1, QTableWidgetItem(str(reportdata['Standardisation'][0]['MostRecentStandard'][0]['Current'])))
        self.page6.PlantPLCTable_5.setItem(2, 1, QTableWidgetItem(str(reportdata['Standardisation'][0]['NumStdPeriodsSinceClear'][0]['Current'])))
        self.page6.PlantPLCTable_5.setItem(3, 1, QTableWidgetItem(str(reportdata['Standardisation'][0]['NumStdPeriodsThisMnth'][0]['Current'])))
        self.page6.PlantPLCTable_5.setItem(0, 2, QTableWidgetItem(str(reportdata['Standardisation'][0]['FirstStandardTime'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_5.setItem(1, 2, QTableWidgetItem(str(reportdata['Standardisation'][0]['MostRecentStandard'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_5.setItem(2, 2, QTableWidgetItem(str(reportdata['Standardisation'][0]['NumStdPeriodsSinceClear'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_5.setItem(3, 2, QTableWidgetItem(str(reportdata['Standardisation'][0]['NumStdPeriodsThisMnth'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_5.setItem(0, 3, QTableWidgetItem(str(reportdata['Standardisation'][0]['FirstStandardTime'][0]['Comment'])))
        self.page6.PlantPLCTable_5.setItem(1, 3, QTableWidgetItem(str(reportdata['Standardisation'][0]['MostRecentStandard'][0]['Comment'])))
        self.page6.PlantPLCTable_5.setItem(2, 3, QTableWidgetItem(str(reportdata['Standardisation'][0]['NumStdPeriodsSinceClear'][0]['Comment'])))
        self.page6.PlantPLCTable_5.setItem(3, 3, QTableWidgetItem(str(reportdata['Standardisation'][0]['NumStdPeriodsThisMnth'][0]['Comment'])))

        # Update Software Versions Table
        Product = QTableWidgetItem(HomeWindow.VersionNumbers.loc[HomeWindow.VersionNumbers["Module"] == "Product", "Version"].to_string(index=False))
        CsSchedule = QTableWidgetItem(HomeWindow.VersionNumbers.loc[HomeWindow.VersionNumbers["Module"] == "CsSchedule", "Version"].to_string(index=False))
        self.page6.PlantPLCTable_6.setItem(0, 0, Product)
        self.page6.PlantPLCTable_6.setItem(1, 0, CsSchedule)
        # Creating the comment item and populating with Blanks
        self.page6.PlantPLCTable_6.setItem(0, 1, QTableWidgetItem(str(reportdata['SoftwareVersions'][0]['Product'][0]['Current'])))
        self.page6.PlantPLCTable_6.setItem(1, 1, QTableWidgetItem(str(reportdata['SoftwareVersions'][0]['CsSchedule'][0]['Current'])))
        self.page6.PlantPLCTable_6.setItem(0, 2, QTableWidgetItem(str(reportdata['SoftwareVersions'][0]['Product'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_6.setItem(1, 2, QTableWidgetItem(str(reportdata['SoftwareVersions'][0]['CsSchedule'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_6.setItem(0, 3, QTableWidgetItem(str(reportdata['SoftwareVersions'][0]['Product'][0]['Comment'])))
        self.page6.PlantPLCTable_6.setItem(1, 3, QTableWidgetItem(str(reportdata['SoftwareVersions'][0]['CsSchedule'][0]['Comment'])))

        # Update Disk Space Table
        DiskSpace = QTableWidgetItem(HomeWindow.VersionNumbers.loc[HomeWindow.VersionNumbers["Module"] == "DiskSpace", "Version"].to_string(index=False))
        PercDiskSpace = QTableWidgetItem(HomeWindow.VersionNumbers.loc[HomeWindow.VersionNumbers["Module"] == "%DiskSpace", "Version"].to_string(index=False))
        self.page6.PlantPLCTable_7.setItem(0, 0, DiskSpace)
        self.page6.PlantPLCTable_7.setItem(1, 0, PercDiskSpace)
        # Creating the comment item and populating with Blanks
        self.page6.PlantPLCTable_7.setItem(0, 1, QTableWidgetItem(str(reportdata['DiskSpaceMem'][0]['DiskSpace'][0]['Current'])))
        self.page6.PlantPLCTable_7.setItem(1, 1, QTableWidgetItem(str(reportdata['DiskSpaceMem'][0]['PercDiskSpace'][0]['Current'])))
        self.page6.PlantPLCTable_7.setItem(0, 2, QTableWidgetItem(str(reportdata['DiskSpaceMem'][0]['DiskSpace'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_7.setItem(1, 2, QTableWidgetItem(str(reportdata['DiskSpaceMem'][0]['PercDiskSpace'][0]['1RepAgo'])))
        self.page6.PlantPLCTable_7.setItem(0, 3, QTableWidgetItem(str(reportdata['DiskSpaceMem'][0]['DiskSpace'][0]['Comment'])))
        self.page6.PlantPLCTable_7.setItem(1, 3, QTableWidgetItem(str(reportdata['DiskSpaceMem'][0]['PercDiskSpace'][0]['Comment'])))

    def NewAnUpdatePg6_1RepAgo(self):
        # **********************************************************************************************************************************************
        # This section defines the filename where we need to pull the relevant PSAreport.xls file from in order to update the 1 Report Ago data
        # **********************************************************************************************************************************************
        anal = self.AnalyserListComboBox.currentText()
        # Switches default folder based on which analyser is selected.
        if anal[0:3] == "C15":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\NG-1500")
        elif anal[0:3] == "C21":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\CS-2100")
        elif anal[0:3] == "CMM":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\CMM-100")
        elif anal[0:3] == "OBA":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\On Belt Analyser")
        elif anal[0:3] == "TBM":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\TBM-210")
        else:
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\TBM-230")

        # ensures the file separators are correct for operating system
        if fname:
            fname = QDir.toNativeSeparators(fname)
            fname2 = glob.glob(fname)
            fname = glob.glob(fname + "\\*Report.xls")
            fname2 = fname2[0]
            fname = fname[0]

        # if user presses cancel, return control to calling function
        if not fname:
            return

        # **************************************************************************************************************************************************
        # This section imports the relevant data into the required data frames
        # **************************************************************************************************************************************************
        print("Filename is: " + str(fname))
        # Import the Analyser Status sheet from the PSAReport.xls file in the relevant folder defined by fname
        StatusColumns1RA = ["Result Name", "Value"]
        AnalyserStatus1RA = pd.read_excel(open(fname, 'rb'), sheet_name="Analyser Status")
        AnalyserStatusdf1RA = pd.DataFrame(AnalyserStatus1RA, columns=StatusColumns1RA)

        # Import the Analyser IO sheet from the PSAReport.xls file in the relevant folder defined by fname
        StatusColumns1RA = ["Parameter Name", "Value"]
        AnalyserIO1RA = pd.read_excel(open(fname, 'rb'), sheet_name="Analyser I_O")
        AnalyserIOdf1RA = pd.DataFrame(AnalyserIO1RA, columns=StatusColumns1RA)

        # Import the Version Numbers sheet from the PSAReport.xls file in the relevant folder defined by fname
        StatusColumns1RA = ["Module", "Version"]
        VersionNumbers1RA = pd.read_excel(open(fname, 'rb'), sheet_name="Version Numbers")
        VersionNumbersdf1RA = pd.DataFrame(VersionNumbers1RA, columns=StatusColumns1RA)
        StdThisMnth = Read_A_17_Standard(fname2)

        # **************************************************************************************************************************************************
        # This section updates the 1 Report ago section of the Page 6 status tables using the dataframes defined earlier
        # **************************************************************************************************************************************************

        # Update PLC Status Table
        BeltRunning = QTableWidgetItem(AnalyserIOdf1RA.loc[AnalyserIOdf1RA["Parameter Name"] == "BeltRunning", "Value"].to_string(index=False))
        ForceAnalyse = QTableWidgetItem(AnalyserIOdf1RA.loc[AnalyserIOdf1RA["Parameter Name"] == "ForceAnalyse", "Value"].to_string(index=False))
        ForceStandardise = QTableWidgetItem(AnalyserIOdf1RA.loc[AnalyserIOdf1RA["Parameter Name"] == "ForceStandardise", "Value"].to_string(index=False))

        self.page6.PlantPLCTable.setItem(0, 1, BeltRunning)
        self.page6.PlantPLCTable.setItem(1, 1, ForceAnalyse)
        self.page6.PlantPLCTable.setItem(2, 1, ForceStandardise)

        # Update Analyser Status Table
        AnalyserOK = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "AnalyserOK", "Value"].to_string(index=False))
        StandardsOK = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "StandardsOK", "Value"].to_string(index=False))
        IOControlOK = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "IOControlOK", "Value"].to_string(index=False))
        SpectraStable = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "SpectraStable", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_2.setItem(0, 1, AnalyserOK)
        self.page6.PlantPLCTable_2.setItem(1, 1, StandardsOK)
        self.page6.PlantPLCTable_2.setItem(2, 1, IOControlOK)
        self.page6.PlantPLCTable_2.setItem(3, 1, SpectraStable)

        # Update PLC Analyser Results
        SystemOK = QTableWidgetItem(AnalyserIOdf1RA.loc[AnalyserIOdf1RA["Parameter Name"] == "SystemRunning", "Value"].to_string(index=False))
        if AnalyserIOdf1RA.loc[AnalyserIOdf1RA["Parameter Name"] == "SourceControlFault", "Value"].to_string(index=False) == "   OK\nFALSE":
            SourceControlFault = QTableWidgetItem("FALSE")
        else:
            SourceControlFault = QTableWidgetItem("TRUE")
        SourceOff = QTableWidgetItem(AnalyserIOdf1RA.loc[AnalyserIOdf1RA["Parameter Name"] == "SourceOffProx", "Value"].to_string(index=False))
        SourceOn = QTableWidgetItem(AnalyserIOdf1RA.loc[AnalyserIOdf1RA["Parameter Name"] == "SourceOnProx", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_3.setItem(0, 1, SystemOK)
        self.page6.PlantPLCTable_3.setItem(1, 1, SourceControlFault)
        self.page6.PlantPLCTable_3.setItem(2, 1, SourceOff)
        self.page6.PlantPLCTable_3.setItem(3, 1, SourceOn)

        # Update Analyser Config Table
        AnalysisPeriod = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "AnalysisPeriod", "Value"].to_string(index=False))
        AnalMinLoadLimit = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "AnalMinLoadLimit", "Value"].to_string(index=False))
        StandardisePeriod = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "StandardisePeriod", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_4.setItem(0, 1, AnalysisPeriod)
        self.page6.PlantPLCTable_4.setItem(1, 1, AnalMinLoadLimit)
        self.page6.PlantPLCTable_4.setItem(2, 1, StandardisePeriod)

        # Update Standardisation Table
        FirstStandard = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "FirstStandardTime", "Value"].to_string(index=False))
        LastStandard = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "LastStandardiseTime", "Value"].to_string(index=False))
        NumStandard = QTableWidgetItem(AnalyserStatusdf1RA.loc[AnalyserStatusdf1RA["Result Name"] == "PeriodCount", "Value"].to_string(index=False))
        NumStandardMonth = QTableWidgetItem(StdThisMnth)
        self.page6.PlantPLCTable_5.setItem(0, 1, FirstStandard)
        self.page6.PlantPLCTable_5.setItem(1, 1, LastStandard)
        self.page6.PlantPLCTable_5.setItem(2, 1, NumStandard)
        self.page6.PlantPLCTable_5.setItem(3, 1, NumStandardMonth)

        # Update Software Versions Table
        Product = QTableWidgetItem(VersionNumbersdf1RA.loc[VersionNumbersdf1RA["Module"] == "Product", "Version"].to_string(index=False))
        CsSchedule = QTableWidgetItem(VersionNumbersdf1RA.loc[VersionNumbersdf1RA["Module"] == "CsSchedule", "Version"].to_string(index=False))
        self.page6.PlantPLCTable_6.setItem(0, 1, Product)
        self.page6.PlantPLCTable_6.setItem(1, 1, CsSchedule)

        # Update Disk Space Table
        DiskSpace = QTableWidgetItem(VersionNumbersdf1RA.loc[VersionNumbersdf1RA["Module"] == "DiskSpace", "Version"].to_string(index=False))
        PercDiskSpace = QTableWidgetItem(VersionNumbersdf1RA.loc[VersionNumbersdf1RA["Module"] == "%DiskSpace", "Version"].to_string(index=False))
        self.page6.PlantPLCTable_7.setItem(0, 1, DiskSpace)
        self.page6.PlantPLCTable_7.setItem(1, 1, PercDiskSpace)

    def NewAnUpdatePg6_2RepAgo(self):
        # **********************************************************************************************************************************************
        # This section defines the filename where we need to pull the relevant PSAreport.xls file from in order to update the 1 Report Ago data
        # **********************************************************************************************************************************************
        anal = self.AnalyserListComboBox.currentText()
        # Switches default folder based on which analyser is selected.
        if anal[0:3] == "C15":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\NG-1500")
        elif anal[0:3] == "C21":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\CS-2100")
        elif anal[0:3] == "CMM":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\CMM-100")
        elif anal[0:3] == "OBA":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\On Belt Analyser")
        elif anal[0:3] == "TBM":
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\TBM-210")
        else:
            fname = QFileDialog.getExistingDirectory(self, "Open PSA File in PSA Folder", "J:\\Client Analysers\\TBM-230")

        # ensures the file separators are correct for operating system
        if fname:
            fname = QDir.toNativeSeparators(fname)
            fname2 = glob.glob(fname)
            fname = glob.glob(fname + "\\*Report.xls")
            fname2 = fname2[0]
            fname = fname[0]

        # if user presses cancel, return control to calling function
        if not fname:
            return

        # **************************************************************************************************************************************************
        # This section imports the relevant data into the required data frames
        # **************************************************************************************************************************************************
        print("Filename is: " + str(fname))
        # Import the Analyser Status sheet from the PSAReport.xls file in the relevant folder defined by fname
        StatusColumns2RA = ["Result Name", "Value"]
        AnalyserStatus2RA = pd.read_excel(open(fname, 'rb'), sheet_name="Analyser Status")
        AnalyserStatusdf2RA = pd.DataFrame(AnalyserStatus2RA, columns=StatusColumns2RA)

        # Import the Analyser IO sheet from the PSAReport.xls file in the relevant folder defined by fname
        StatusColumns2RA = ["Parameter Name", "Value"]
        AnalyserIO2RA = pd.read_excel(open(fname, 'rb'), sheet_name="Analyser I_O")
        AnalyserIOdf2RA = pd.DataFrame(AnalyserIO2RA, columns=StatusColumns2RA)

        # Import the Version Numbers sheet from the PSAReport.xls file in the relevant folder defined by fname
        StatusColumns2RA = ["Module", "Version"]
        VersionNumbers2RA = pd.read_excel(open(fname, 'rb'), sheet_name="Version Numbers")
        VersionNumbersdf2RA = pd.DataFrame(VersionNumbers2RA, columns=StatusColumns2RA)
        StdThisMnth = Read_A_17_Standard(fname2)

        # **************************************************************************************************************************************************
        # This section updates the 1 Report ago section of the Page 6 status tables using the dataframes defined earlier
        # **************************************************************************************************************************************************

        # Update PLC Status Table
        BeltRunning = QTableWidgetItem(AnalyserIOdf2RA.loc[AnalyserIOdf2RA["Parameter Name"] == "BeltRunning", "Value"].to_string(index=False))
        ForceAnalyse = QTableWidgetItem(AnalyserIOdf2RA.loc[AnalyserIOdf2RA["Parameter Name"] == "ForceAnalyse", "Value"].to_string(index=False))
        ForceStandardise = QTableWidgetItem(AnalyserIOdf2RA.loc[AnalyserIOdf2RA["Parameter Name"] == "ForceStandardise", "Value"].to_string(index=False))

        self.page6.PlantPLCTable.setItem(0, 2, BeltRunning)
        self.page6.PlantPLCTable.setItem(1, 2, ForceAnalyse)
        self.page6.PlantPLCTable.setItem(2, 2, ForceStandardise)

        # Update Analyser Status Table
        AnalyserOK = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "AnalyserOK", "Value"].to_string(index=False))
        StandardsOK = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "StandardsOK", "Value"].to_string(index=False))
        IOControlOK = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "IOControlOK", "Value"].to_string(index=False))
        SpectraStable = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "SpectraStable", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_2.setItem(0, 2, AnalyserOK)
        self.page6.PlantPLCTable_2.setItem(1, 2, StandardsOK)
        self.page6.PlantPLCTable_2.setItem(2, 2, IOControlOK)
        self.page6.PlantPLCTable_2.setItem(3, 2, SpectraStable)

        # Update PLC Analyser Results
        SystemOK = QTableWidgetItem(AnalyserIOdf2RA.loc[AnalyserIOdf2RA["Parameter Name"] == "SystemRunning", "Value"].to_string(index=False))
        if AnalyserIOdf2RA.loc[AnalyserIOdf2RA["Parameter Name"] == "SourceControlFault", "Value"].to_string(index=False) == "   OK\nFALSE":
            SourceControlFault = QTableWidgetItem("FALSE")
        else:
            SourceControlFault = QTableWidgetItem("TRUE")
        SourceOff = QTableWidgetItem(AnalyserIOdf2RA.loc[AnalyserIOdf2RA["Parameter Name"] == "SourceOffProx", "Value"].to_string(index=False))
        SourceOn = QTableWidgetItem(AnalyserIOdf2RA.loc[AnalyserIOdf2RA["Parameter Name"] == "SourceOnProx", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_3.setItem(0, 2, SystemOK)
        self.page6.PlantPLCTable_3.setItem(1, 2, SourceControlFault)
        self.page6.PlantPLCTable_3.setItem(2, 2, SourceOff)
        self.page6.PlantPLCTable_3.setItem(3, 2, SourceOn)

        # Update Analyser Config Table
        AnalysisPeriod = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "AnalysisPeriod", "Value"].to_string(index=False))
        AnalMinLoadLimit = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "AnalMinLoadLimit", "Value"].to_string(index=False))
        StandardisePeriod = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "StandardisePeriod", "Value"].to_string(index=False))
        self.page6.PlantPLCTable_4.setItem(0, 2, AnalysisPeriod)
        self.page6.PlantPLCTable_4.setItem(1, 2, AnalMinLoadLimit)
        self.page6.PlantPLCTable_4.setItem(2, 2, StandardisePeriod)

        # Update Standardisation Table
        FirstStandard = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "FirstStandardTime", "Value"].to_string(index=False))
        LastStandard = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "LastStandardiseTime", "Value"].to_string(index=False))
        NumStandard = QTableWidgetItem(AnalyserStatusdf2RA.loc[AnalyserStatusdf2RA["Result Name"] == "PeriodCount", "Value"].to_string(index=False))
        NumStandardMonth = QTableWidgetItem(StdThisMnth)
        self.page6.PlantPLCTable_5.setItem(0, 2, FirstStandard)
        self.page6.PlantPLCTable_5.setItem(1, 2, LastStandard)
        self.page6.PlantPLCTable_5.setItem(2, 2, NumStandard)
        self.page6.PlantPLCTable_5.setItem(3, 2, NumStandardMonth)

        # Update Software Versions Table
        Product = QTableWidgetItem(VersionNumbersdf2RA.loc[VersionNumbersdf2RA["Module"] == "Product", "Version"].to_string(index=False))
        CsSchedule = QTableWidgetItem(VersionNumbersdf2RA.loc[VersionNumbersdf2RA["Module"] == "CsSchedule", "Version"].to_string(index=False))
        self.page6.PlantPLCTable_6.setItem(0, 2, Product)
        self.page6.PlantPLCTable_6.setItem(1, 2, CsSchedule)

        # Update Disk Space Table
        DiskSpace = QTableWidgetItem(VersionNumbersdf2RA.loc[VersionNumbersdf2RA["Module"] == "DiskSpace", "Version"].to_string(index=False))
        PercDiskSpace = QTableWidgetItem(VersionNumbersdf2RA.loc[VersionNumbersdf2RA["Module"] == "%DiskSpace", "Version"].to_string(index=False))
        self.page6.PlantPLCTable_7.setItem(0, 2, DiskSpace)
        self.page6.PlantPLCTable_7.setItem(1, 2, PercDiskSpace)

    def UpdateFigs(self):
        self.page2.label_3.setPixmap(QtGui.QPixmap(HomeWindow.resourcepath + "\\Detector_Stability.png"))
        self.page3.Temps.setPixmap(QtGui.QPixmap(HomeWindow.resourcepath + "\\Temperatures.png"))
        self.page3.label_4.setPixmap(QtGui.QPixmap(HomeWindow.resourcepath + "\\Daily_Tonnes.png"))
        self.page4.Results1.setPixmap(QtGui.QPixmap(HomeWindow.resourcepath + "\\Results1.png"))
        self.page5.Results2.setPixmap(QtGui.QPixmap(HomeWindow.resourcepath + "\\Results2.png"))

    def UpdateDetStab(self):
        # calculate the number of detecotrs
        Dets = HomeWindow.PeakControl.shape[1] - 2

        # clear the table each time the function is called.
        for i in range(16):
            self.page2.tableWidget.setItem(0, i, self.ClearTable())
            self.page2.tableWidget.setItem(1, i, self.ClearTable())

        x = 0
        # update the detecotr status table based on the totoal number of detectors.
        while x < Dets:
            # read in the detector disabled status
            disabled = HomeWindow.PeakControl.loc[HomeWindow.PeakControl["Result Name"] == "DetectorDisabled", "Detector " + str(x + 1)].to_string(index=False)
            # read in the detector Stability status
            stablity = HomeWindow.PeakControl.iloc[21, x + 1]

            # Choose which colour and wording to update the Detector enabled row of detecotr stability table with
            if disabled == "FALSE":
                value = self.EnabledYes()
            else:
                value = self.EnabledNo()
            # update the Enabled / Disabled row of the detector stability table with the wording chosen above
            self.page2.tableWidget.setItem(0, x, value)

            # Choose which colour and wording to update the Stability row of detecotr stability table with
            if stablity == "Stable":
                value = self.EnabledYes()
            else:
                value = self.EnabledNo()
            # update the stability row of the detector stability table with the wording chosen above
            self.page2.tableWidget.setItem(1, x, value)
            # incrment the loop counter
            x += 1

    def RepPeriod(self):
        dates = HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "LastPsaReportTime", "Value"]  # get the date of the last psa report generation
        dates1 = pd.to_datetime(dates, yearfirst=True)  # setup dataframe as a date time series in the correct format
        Period = dates1.dt.month_name().to_string(index=False) + " " + dates1.dt.year.to_string(index=False)  # concatenate the month name and year integer into a string called period using the to_string method to eliminate indexs
        return Period  # return Period to main

    def RepEmail(self, region):

        if region == "Asia":
            email = "service.asia@scantech.com.au"
        elif region == "Australia":
            email = "service.australia@scantech.com.au"
        elif region == "China":
            email = "service.asia@scantech.com.au"
        elif region == "Europe":
            email = "service.europe@scantech.com.au"
        elif region == "India":
            email = "service.india@scantech.com.au"
        elif region == "Middle East & India":
            email = "service.middleeast@scantech.com.au"
        elif region == "New Zealand":
            email = "service.newzealand@scantech.com.au"
        elif region == "North Africa":
            email = "service.northafrica@scantech.com.au"
        elif region == "South Africa":
            email = "service.southafrica@scantech.com.au"
        elif region == "Latin America":
            email = "service.southamerica@scantech.com.au"
        elif region == "USA & Canada":
            email = "service.usa@scantech.com.au"
        else:
            email = "service.fsu@scantech.com.au"

        return email

    def StdDate(self):
        stddate = HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "LastStandardiseTime", "Value"]  # get the date of the last psa report generation
        stddate = stddate.to_string(index=False)  # setup dataframe as a date time series in the correct format
        stddate = datetime.datetime.strptime(stddate, "%a %y/%m/%d %H:%M")
        tdate = datetime.datetime.today()
        difference = (tdate.year - stddate.year) * 12 + (tdate.month - stddate.month)
        if difference > 6:
            output = "No"
        else:
            output = "Yes"

        return output

    def DiskSpaceOK(self):
        diskspace = HomeWindow.VersionNumbers.loc[HomeWindow.VersionNumbers["Module"] == "%DiskSpace", "Version"]
        diskspace = diskspace.to_string(index=False)
        diskspace = float(diskspace.replace("%", ""))
        if diskspace > 10.00:
            diskspaceok = "Yes"
        else:
            diskspaceok = "No"

        return diskspaceok

    def openReport(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self, "Choose Report File to Open", "", "csv (*.csv)",
                                                  options=options)
        print(fileName)

    def popanalysers(self):
        configdatafname = "J:\\Client Analysers\\Analyser PSA Report\\PSA Report Config data.xlsx"  # generates the checklist filename based on the year of the period of the report being produced. This will deal with the January / december issue
        configdata = openpyxl.load_workbook(configdatafname)  # Load config data from Config Data spreadsheet
        andetails = configdata["AnalyserDetails"]  # Select Analyser Details sheet
        maxrow = andetails.max_row  # Determine max row of document
        for row in range(1, maxrow + 1):  # For loop for searching through the Analyser Details sheet to determine the elemental result names to be plotted.
            ancol = "{}{}".format("A", row)  # this sets up the analyser cell reference to be used.
            reqcol = "{}{}".format("G", row)  # this sets up the psa report required reference cell
            repreq = andetails[reqcol].value
            if repreq == "Yes":
                x = andetails[ancol].value
                self.AnalyserListComboBox.addItem(str(x))
            else:
                pass

    def UpdatePSAChecklist(self):  # this must only be called after the JSON has been updated since this function uses that JSON to update the PSA Checklist
        with open(HomeWindow.jsonoutfname) as f:  # reload the json file to use in the function
            jsondata = json.load(f)
        if HomeWindow.blankrep == 1:
            periodd = datetime.date.today() + relativedelta(months=-1)  # Subtract a month from today's date as the report is for the previous month
            period = periodd.strftime("%B %Y")  # Report Period string - Month & Year
            year = periodd.strftime("%Y")

        else:
            dates = HomeWindow.AnalyserStatus.loc[HomeWindow.AnalyserStatus["Result Name"] == "LastPsaReportTime", "Value"]  # get the date of the last psa report generation. This gives the period of the report.
            dates1 = pd.to_datetime(dates, yearfirst=True)  # setup dataframe as a date time series in the correct format
            period = dates1.dt.month_name().to_string(index=False) + " " + dates1.dt.year.to_string(index=False)  # concatenate the month name and year integer into a string called period using the to_string method to eliminate indexs
            year = dates1.dt.year.to_string(index=False)  # get the year of the report

        checklistfname = "C:\\PSAGen\\PSA Report Checklist " + year + ".xlsx"  # generates the checklist filename based on the year of the period of the report being produced. This will deal with the January / december issue
        checklist = openpyxl.load_workbook(checklistfname)  # load the relevant checklist
        psachecklist = checklist["PSA Checklist"]  # load the relevant sheet from the checklist
        analyser = jsondata["Summary"][0]["AnalyserNo"]  # pull the analyser in question from the json
        analyser = analyser.replace("-", "")  # remove the hyphens from the analyser name
        maxrow = psachecklist.max_row  # calculate the max number of rows in the checklist
        for row in range(1, maxrow + 1):  # For loop to search for the relevant analyser to update in the checklist
            job = "{}{}".format("A", row)  # this sets up the cell reference in the PSA Checklist Excel document that we are going to search through.
            x = psachecklist[job].value  # pull the relevant analyser from the checklist
            if x == analyser:  # if statement which generates the correct cell references for the function to update.
                if period == "January " + year:
                    c1 = "{}{}".format("E", row)
                    c2 = "{}{}".format("F", row)
                elif period == "February " + year:
                    c1 = "{}{}".format("J", row)
                    c2 = "{}{}".format("K", row)
                elif period == "March " + year:
                    c1 = "{}{}".format("O", row)
                    c2 = "{}{}".format("P", row)
                elif period == "April " + year:
                    c1 = "{}{}".format("T", row)
                    c2 = "{}{}".format("U", row)
                elif period == "May " + year:
                    c1 = "{}{}".format("Y", row)
                    c2 = "{}{}".format("Z", row)
                elif period == "June " + year:
                    c1 = "{}{}".format("AD", row)
                    c2 = "{}{}".format("AE", row)
                elif period == "July " + year:
                    c1 = "{}{}".format("AI", row)
                    c2 = "{}{}".format("AJ", row)
                elif period == "August " + year:
                    c1 = "{}{}".format("AN", row)
                    c2 = "{}{}".format("AO", row)
                elif period == "September " + year:
                    c1 = "{}{}".format("AS", row)
                    c2 = "{}{}".format("AT", row)
                elif period == "October " + year:
                    c1 = "{}{}".format("AX", row)
                    c2 = "{}{}".format("AY", row)
                elif period == "November " + year:
                    c1 = "{}{}".format("BC", row)
                    c2 = "{}{}".format("BD", row)
                else:
                    c1 = "{}{}".format("BH", row)
                    c2 = "{}{}".format("BI", row)
                if HomeWindow.blankrep == 1:
                    psachecklist[c1].value = "Blank"
                else:
                    psachecklist[c1].value = jsondata["Summary"][0]["ReportDate"]
                psachecklist[c2].value = jsondata["Summary"][0]["ReportDate"]
                break
            else:
                pass

        openpyxl.Workbook.save(checklist, checklistfname)  # save and overwrite the updated checklist
        print("PSA Checklist Updated")

    def IssueBlank(self):
        # This function generates a blank PSA report when required.
        self.ImportReportPB.setEnabled(False)  # Disable Import Report Pushbutton
        HomeWindow.blankrep = 1  # Set the Blank report flag
        configdatafname = "J:\\Client Analysers\\Analyser PSA Report\\PSA Report Config data.xlsx"  # generates the checklist filename based on the year of the period of the report being produced. This will deal with the January / december issue
        configdata = openpyxl.load_workbook(configdatafname)  # Load config data from Config Data spreadsheet
        andetails = configdata["AnalyserDetails"]  # Select Analyser Details sheet
        maxrow = andetails.max_row  # Determine max row of document
        analyser = self.AnalyserListComboBox.currentText()  # get the relevant analyser from the GUI Combobox
        for row in range(1, maxrow + 1):  # For loop for searching through the Analyser Details sheet to determine the elemental result names to be plotted.
            ancol = "{}{}".format("A", row)  # this sets up the analyser cell reference to be used.
            appcol = "{}{}".format("D", row)  # this sets up the application reference cell
            engcol = "{}{}".format("C", row)  # this sets up the service engineer reference cell
            regcol = "{}{}".format("F", row)  # this sets up the service engineer reference cell
            custcol = "{}{}".format("B", row)  # this sets up the service engineer reference cell
            blankanal = andetails[ancol].value  # this puts the analyser which is getting a blank report into a variable
            if blankanal == analyser:  # if statement to fine the correct analyser in the list
                app = andetails[appcol].value  # determine the application fo the variable
                serveng = andetails[engcol].value  # determine the relevant service engineer
                region = andetails[regcol].value  # determine the relevant region
                regemail = self.RepEmail(region)  # determine the relevant regional email address
                repdate = datetime.date.today().strftime("%d %B %Y")  # get Today's date and convert to a string
                period = datetime.date.today() + relativedelta(months=-1)  # Subtract a month from today's date as the report is for the previous month
                shortm = period.strftime("%m")  # Report Period string - Month Short format
                shorty = period.strftime("%y")  # Report Period string - Year Short Format
                period = period.strftime("%B %Y")  # Report Period string - Month & Year
                cust = andetails[custcol].value  # get customer details
                reportdata['Summary'][0]['SiteName'] = cust  # populate the
                reportdata['Summary'][0]['ReportDate'] = repdate
                reportdata['Summary'][0]['AnalyserNo'] = analyser
                reportdata['Summary'][0]['ServEng'] = serveng
                reportdata['Summary'][0]['Application'] = app
                reportdata['Summary'][0]['Period'] = period
                reportdata['Summary'][0]['Email'] = regemail
                reportdata['Summary'][0]['AnalyserOpCorrect'] = "Blank"
                reportdata['ActionTaken'][0]['Action1'][0]['Date'] = repdate
                reportdata['ActionTaken'][0]['Action1'][0]['Action'] = "Unsuccessful attempt to compile report"
                reportdata['ActionTaken'][0]['Action1'][0]['Description'] = "No PSA Data Available due to no remote connection or analyser offline"
                reportdata['ActionRequired'][0]['ActionReq1'][0]['Action'] = "Attempt to resolve why analyser or remote connection is offline"
                reportdata['ActionRequired'][0]['ActionReq1'][0]['ByWhom'] = "Site\\Scantech"
                reportdata['ActionRequired'][0]['ActionReq1'][0]['ByWhen'] = "As Soon As Possible"
                HomeWindow.psadatafname = self.GetFolder("Select Folder to save blank JSON")  # Get the filename of the folder where the JSON will be dumped.
                HomeWindow.jsonname = "Blank " + str(analyser) + " PSA Report " + shorty + shortm + ".json"  # setting the filename which will be used to output the JSON file
                HomeWindow.jsonoutfname = HomeWindow.psadatafname + "\\" + HomeWindow.jsonname
                with open(HomeWindow.jsonoutfname, 'w') as file:
                    json.dump(reportdata, file, indent=1)
                break
            else:
                pass
        self.ExportReportPB.setEnabled(True)  # Enable the export JSON pushbutton.

    def hide_pages(self):
        self.page1.hide()
        self.page2.hide()
        self.page3.hide()
        self.page4.hide()
        self.page5.hide()
        self.page6.hide()

    def show_page_1(self):
        self.hide_pages()
        self.page1.show()

    def show_page_2(self):
        self.hide_pages()
        self.page2.show()

    def show_page_3(self):
        self.hide_pages()
        self.page3.show()

    def show_page_4(self):
        self.hide_pages()
        self.page4.show()

    def show_page_5(self):
        self.hide_pages()
        self.page5.show()

    def show_page_6(self):
        self.hide_pages()
        self.page6.show()


class PSA_pg1(QMainWindow, Ui_PSAPage1):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)


class PSA_pg2(QMainWindow, Ui_PSAPage2):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)


class PSA_pg3(QMainWindow, Ui_PSAPage3):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)


class PSA_pg4(QMainWindow, Ui_PSAPage4):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)


class PSA_pg5(QMainWindow, Ui_PSAPage5):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)


class PSA_pg6(QMainWindow, Ui_PSAPage6):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)


app = QApplication(sys.argv)
window = HomeWindow()
window.show()

app.exec()
